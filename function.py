import json
import os
from openpyxl import load_workbook
import csv

def read_json():
    # 读json文件，temp.json文件用于预览，setting.json文件用于UI设计，否则则用默认设置
    if os.path.exists("temp.json"):
        with open('temp.json', 'r', encoding="utf-8") as json_file:
            data = json.load(json_file)
        return data
    elif os.path.exists("setting.json"):
        with open('setting.json', 'r', encoding="utf-8") as json_file:
            data = json.load(json_file)
        return data
    else:
        data = {
            "main": {
                "window_background_color": "#f7fed5", "label_font_size": 20, "title_font_size": 30, "title_background_color": "#94d2ef",
                "button_font_size": 30, "button_padding": 10, "button_border_radius": 10, "button_background_color": "#e7e7e7",
                "button_pressed_color": "#145214", "filepath_font_size": 18
            },
            "slide": {
                "window_background_color": "#e7e7e7", "label_font_size": 20, "label_background_color": "#ffffff", "answer_font_size": 20,
                "answer_background_color": "#000000", "answer_background_press_color": "#ffffff", "button_font_size": 18, "button_padding": 10,
                "button_border_radius": 10, "button_background_color": "#ffffff", "button_pressed_color": "#00ffff"
            },
            "single":{
                "window_background_color": "#e7e7e7", "label_font_size": 20, "label_background_color": "#ffffff",
                "answer_font_size": 20, "answer_background_color": "#ffffff", "button_font_size": 18, "button_padding": 10,
                "button_border_radius": 10, "button_background_color": "#ffffff", "button_pressed_color": "#00ffff",
                "button_right_color": "#00ff00", "button_wrong_color": "#ff0000", "non_repeat_check": True
            }
        }
        return data

def get_file_path():
    # 获取excel文件路径
    if os.path.exists("information.txt"):
        with open("information.txt", "r", encoding="utf-8") as f:
            fpath = f.read().strip()
    else:
        dirpath = os.getcwd()
        fpath = os.path.join(dirpath, "Templete.xlsx")
    return fpath

def read_random_data_from_excel(fpath):
    # 读取excel文件
    workbook = load_workbook(filename=fpath, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(min_row=2, values_only=True)
    wordlist = [row for row in rows]
    return wordlist

def read_random_data_from_csv(fpath):
    # 读取csv文件
    with open(fpath, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        wordlist = [row for row in reader]
    return wordlist

if __name__ == '__main__':
    print()